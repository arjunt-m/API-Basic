import os
import json
import openpyxl


class Utils:

    @staticmethod
    def add_data_from_excel(sheet_name, testcase_id=None, row_number=None):
        dir_root = os.path.dirname(os.path.abspath(__file__))
        project_root = os.path.dirname(dir_root)
        json_path = os.path.join(
            project_root,
            "Requests",
            "Post_Request_Excel.json"
        )
        with open(json_path, "r") as file:
            payload_template = json.load(file)
        excel_path = os.path.join(
            project_root,
            "Requests",
            "Request.xlsx"
        )
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook[sheet_name]
        max_row = sheet.max_row
        max_col = sheet.max_column
        headers = []

        # Read headers
        for col in range(1, max_col + 1):
            headers.append(sheet.cell(row=1, column=col).value)

        data_list = []

        # Decide which rows to read
        if row_number:
            rows = [row_number]
        else:
            rows = range(2, max_row + 1)

        for row in rows:
            new_payload = payload_template.copy()
            current_testcase = None
            for col in range(1, max_col + 1):
                key = headers[col - 1]
                value = sheet.cell(row=row, column=col).value
                # Convert datetime to string
                if hasattr(value, "strftime"):
                    value = value.strftime("%Y-%m-%d")
                # Capture TestCaseID
                if key == "TestCaseID":
                    current_testcase = value
                    continue
                new_payload[key] = value
            # Filter by TestCaseID if provided
            if testcase_id:
                if current_testcase == testcase_id:
                    data_list.append(new_payload)
            else:
                data_list.append(new_payload)
        workbook.close()
        return data_list